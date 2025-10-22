import streamlit as st
import gurobipy as gp
from gurobipy import GRB
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import numpy as np
import json
import os
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule



# ---------------------- Functions Start ----------------------
def validate_demand_json(demand_config):
    if "periods" not in demand_config or not isinstance(demand_config["periods"], list):
        return False
    for period in demand_config["periods"]:
        if not all(k in period for k in ["name", "hours", "demand_mw"]):
            return False
    return True

def validate_plants_json(plants_config):
    for key in ["thermal_plants", "hydro_plants"]:
        if key not in plants_config or not isinstance(plants_config[key], list):
            return False
        for plant in plants_config[key]:
            if not all(k in plant for k in ["id", "name", "default", "ranges"]):
                return False
    return True

# Load configuration files
@st.cache_data
def load_config(uploaded_demand=None, uploaded_plants=None):
    # Default demand configuration
    default_demand = {
        "periods": [
            {"name": "0h-6h", "hours": 6, "demand_mw": 15000},
            {"name": "6h-9h", "hours": 3, "demand_mw": 30000},
            {"name": "9h-15h", "hours": 6, "demand_mw": 25000},
            {"name": "15h-18h", "hours": 3, "demand_mw": 40000},
            {"name": "18h-24h", "hours": 6, "demand_mw": 27000}
        ]
    }
    
    # Default plant types configuration
    default_plants = {
        "thermal_plants": [
            {
                "id": "A",
                "name": "Type A",
                "default": {"amount": 12, "p_min": 850, "p_max": 2000, "c_base": 1000, "c_mwh": 2.0, "c_start": 2000},
                "ranges": {"amount": [1, 20], "p_min": [100, 2000, 50], "p_max": [500, 5000, 50], 
                          "c_base": [0, 5000, 50], "c_mwh": [0.0, 5.0, 0.1], "c_start": [0, 3000, 50]}
            },
            {
                "id": "B",
                "name": "Type B",
                "default": {"amount": 10, "p_min": 1250, "p_max": 1750, "c_base": 2600, "c_mwh": 1.3, "c_start": 1000},
                "ranges": {"amount": [1, 20], "p_min": [100, 2000, 50], "p_max": [500, 5000, 50], 
                          "c_base": [0, 5000, 50], "c_mwh": [0.0, 5.0, 0.1], "c_start": [0, 3000, 50]}
            },
            {
                "id": "C",
                "name": "Type C",
                "default": {"amount": 5, "p_min": 1500, "p_max": 4000, "c_base": 3000, "c_mwh": 3.0, "c_start": 500},
                "ranges": {"amount": [1, 20], "p_min": [100, 2000, 50], "p_max": [500, 5000, 50], 
                          "c_base": [0, 5000, 50], "c_mwh": [0.0, 5.0, 0.1], "c_start": [0, 3000, 50]}
            }
        ],
        "hydro_plants": [
            {
                "id": "H1",
                "name": "Hydro 1",
                "default": {"amount": 1, "p_min": 900, "p_max": 900, "c_base": 90, "c_mwh": 0, "c_start": 1500, "v_abai": 0.31},
                "ranges": {"amount": [0, 2], "p_min": [500, 2000, 50], "p_max": [500, 2000, 50], 
                          "c_base": [0, 500, 10], "c_start": [0, 2000, 50], "v_abai": [0.0, 1.0, 0.01]}
            },
            {
                "id": "H2",
                "name": "Hydro 2",
                "default": {"amount": 1, "p_min": 1400, "p_max": 1400, "c_base": 150, "c_mwh": 0, "c_start": 1200, "v_abai": 0.47},
                "ranges": {"amount": [0, 2], "p_min": [500, 2000, 50], "p_max": [500, 2000, 50], 
                          "c_base": [0, 500, 10], "c_start": [0, 2000, 50], "v_abai": [0.0, 1.0, 0.01]}
            }
        ]
    }
    
        # Load demand
    if uploaded_demand:
        try:
            demand_config = json.load(uploaded_demand)
            if not validate_demand_json(demand_config):
                st.warning("❌ Uploaded demand JSON is invalid. Using default demand configuration.")
                demand_config = default_demand
        except:
            st.warning("❌ Could not read uploaded demand JSON. Using default demand configuration.")
            demand_config = default_demand
    elif os.path.exists('demand_config.json'):
        with open('demand_config.json', 'r') as f:
            demand_config = json.load(f)
    else:
        demand_config = default_demand

    # Load plants
    if uploaded_plants:
        try:
            plants_config = json.load(uploaded_plants)
            if not validate_plants_json(plants_config):
                st.warning("❌ Uploaded plants JSON is invalid. Using default plants configuration.")
                plants_config = default_plants
        except:
            st.warning("❌ Could not read uploaded plants JSON. Using default plants configuration.")
            plants_config = default_plants
    elif os.path.exists('plants_config.json'):
        with open('plants_config.json', 'r') as f:
            plants_config = json.load(f)
    else:
        plants_config = default_plants

    return demand_config, plants_config


def create_model(demand_df, plants, reserve_pct, cyclic, include_pumping):
    # Create model
    model = gp.Model("Electricity_Production")
    model.setParam('OutputFlag', 0)
    model.ModelSense = GRB.MINIMIZE
    
    # Decision variables
    N, X, P = {}, {}, {}
    for t_name in plants.keys():
        for p in range(1, len(demand_df) + 1):
            N[f"{t_name},{p}"] = model.addVar(lb=0, vtype=GRB.INTEGER, name=f"N_{t_name}_{p}")
            X[f"{t_name},{p}"] = model.addVar(lb=0, vtype=GRB.INTEGER, name=f"X_{t_name}_{p}")
            P[f"{t_name},{p}"] = model.addVar(lb=0, vtype=GRB.CONTINUOUS, name=f"P_{t_name}_{p}")
    
    # Objective function
    cost = 0
    for p in range(1, len(demand_df) + 1):
        hours = demand_df.iloc[p-1]['Hours']
        for t_name, t_data in plants.items():
            cost += (P[f"{t_name},{p}"] - N[f"{t_name},{p}"] * t_data['pMin']) * t_data['cMwh'] * hours
            cost += N[f"{t_name},{p}"] * t_data['cBase'] * hours
            cost += X[f"{t_name},{p}"] * t_data['cStart']
    model.setObjective(cost)
    
    # Constraints
    for p in range(1, len(demand_df) + 1):
        demand_val = demand_df.iloc[p-1]['Demand (MW)']
        hours = demand_df.iloc[p-1]['Hours']
        
        # Demand
        model.addConstr(sum(P[f"{t},{p}"] for t in plants.keys()) >= demand_val, name=f"Demand_{p}")
        
        # Reserve
        model.addConstr(sum(N[f"{t},{p}"] * plants[t]['pMax'] - P[f"{t},{p}"] for t in plants.keys()) 
                        >= reserve_pct * demand_val, name=f"Reserve_{p}")
        
        for t_name, t_data in plants.items():
            # Power limits
            model.addConstr(N[f"{t_name},{p}"] * t_data['pMin'] <= P[f"{t_name},{p}"], name=f"Pmin_{t_name}_{p}")
            model.addConstr(P[f"{t_name},{p}"] <= N[f"{t_name},{p}"] * t_data['pMax'], name=f"Pmax_{t_name}_{p}")
            
            # Capacity
            model.addConstr(N[f"{t_name},{p}"] <= t_data['amount'], name=f"Cap_{t_name}_{p}")
            
            # Startup logic
            if p == 1:
                prev_N = N[f"{t_name},{len(demand_df)}"] if cyclic else 0
            else:
                prev_N = N[f"{t_name},{p-1}"]
            model.addConstr(N[f"{t_name},{p}"] <= X[f"{t_name},{p}"] + prev_N, name=f"Start_{t_name}_{p}")
    
    # Pumping constraint
    if include_pumping:
        water_used = sum(3000 * plants[t]['vAbai'] * N[f"{t},{p}"] * demand_df.iloc[p-1]['Hours'] 
                        for t in plants.keys() for p in range(1, len(demand_df) + 1))
        total_production = sum(P[f"{t},{p}"] * demand_df.iloc[p-1]['Hours'] 
                                for t in plants.keys() for p in range(1, len(demand_df) + 1))
        total_demand_energy = sum(demand_df.iloc[p-1]['Demand (MW)'] * demand_df.iloc[p-1]['Hours'] 
                                for p in range(len(demand_df)))
        model.addConstr(total_production >= total_demand_energy + water_used, name="Pumping")

    return model, N, X, P

def optimize(demand_df, plants, reserve_pct, cyclic, include_pumping):
    try:
        model, N, X, P = create_model(demand_df, plants, reserve_pct, cyclic, include_pumping)
        model.update()
        model.optimize()

        # Gather runtime stats similar to tableau_final
        if model.status == GRB.OPTIMAL:
            meilleure_valeur = model.ObjVal
            gap = model.MIPGap
            stats = {
                "Model Name": model.model_name,
                "Nombre de variables": model.NumVars,
                "Nombre de contraintes": model.NumConstrs,
                "Nombre de non-zeros": model.NumNZs,
                "Nombre de nœuds explorés": model.NodeCount,
                "Durée de résolution (s)": model.Runtime,
                "Meilleure valeur objective trouvée (kilo euro)": round(meilleure_valeur/1000),
                "Gap d'optimalité": f"{gap*100:.2f}"
            }
        else:
            stats = {
                "Model Name": model.model_name,
                "Nombre de variables": model.NumVars,
                "Nombre de contraintes": model.NumConstrs,
                "Nombre de non-zeros": model.NumNZs,
                "Nombre de nœuds explorés": None,
                "Durée de résolution (s)": model.Runtime,
                "Meilleure valeur objective trouvée (kilo euro)": None,
                "Gap d'optimalité": None
            }

        # Store results
        # Store results safely
        results_data = []
        for p in range(1, len(demand_df) + 1):
            period_name = demand_df.iloc[p-1]['Period']
            demand_val = demand_df.iloc[p-1]['Demand (MW)']
            for t_name in plants.keys():
                # Safe access: only get .X if model is optimal
                n_val = N[f"{t_name},{p}"].X if model.status == GRB.OPTIMAL else None
                x_val = X[f"{t_name},{p}"].X if model.status == GRB.OPTIMAL else None
                p_val = P[f"{t_name},{p}"].X if model.status == GRB.OPTIMAL else None
                
                results_data.append({
                    'Period': period_name,
                    'Period_Num': p,
                    'Plant': t_name,
                    'N_active': n_val,
                    'X_started': x_val,
                    'P_power': p_val,
                    'Demand': demand_val
                })


        st.session_state.results = {
            'df': pd.DataFrame(results_data),
            'objective': model.ObjVal if model.status == GRB.OPTIMAL else None,
            'status': 'Optimal' if model.status == GRB.OPTIMAL else ('Infeasible' if model.status == GRB.INFEASIBLE else str(model.status)),
            'runtime': model.Runtime,
            'gap': model.MIPGap if model.status == GRB.OPTIMAL else None,
            'num_vars': model.NumVars,
            'num_constrs': model.NumConstrs,
            'nodes_explored': model.NodeCount,
            'model_name': model.model_name
        }
        st.session_state.optimized = True
        st.session_state.model_stats = stats  # <-- store full stats
        st.rerun()

        # Handle infeasible
        if model.status == GRB.INFEASIBLE:
            model.computeIIS()
            infeasible_constrs = [c.ConstrName for c in model.getConstrs() if c.IISConstr]
            st.session_state.results['constraints'] = infeasible_constrs

    except Exception as e:
        import traceback
        st.error(f"❌ Error during optimization: {str(e)}")
        st.code(traceback.format_exc())

def footnote(img_path, name, year, version, link):
    """Adiciona um rodapé estiloso com imagem e texto"""
    import base64
    import streamlit as st

    with open(img_path, "rb") as f:
        data = f.read()
    b64 = base64.b64encode(data).decode()

    st.markdown("---")
    st.markdown(f"""
    <style>
    .footer-container {{
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 20px;
        margin-top: 20px;
        padding: 10px 20px;
        border-radius: 12px;
        transition: box-shadow 0.3s ease, transform 0.3s ease, background-color 0.3s ease;
        font-family: 'Comic Sans MS', 'Cursive', sans-serif;
        background-color: transparent;
    }}

    .footer-container:hover {{
        background-color: #FFFFFF; /* white background on hover */
        box-shadow: -8px 0 20px rgba(0,0,0,0.3);
        transform: translateX(-2px);
    }}

    .footer-container a {{
        text-decoration: none;
    }}

    .footer-text {{
        display: flex;
        flex-direction: column;
        justify-content: center;
        font-size: 1rem;
        color: #333333; /* always dark gray text */
        font-weight: bold;
    }}
    </style>

    <div class="footer-container">
        <a href="{link}" target="_blank">
            <img src="data:image/png;base64,{b64}" width="150" style="border-radius:50%;">
        </a>
        <div class="footer-text">
            <span>{name}</span>
            <span>{year}</span>
            <span>v{version}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ---------------------- Functions End ----------------------

# ---------------------- Load configurations Start ----------------------
# Page configuration
st.set_page_config(page_title="Electricity Production Optimizer", layout="wide", page_icon="⚡")

# Initialize session state
if 'optimized' not in st.session_state:
    st.session_state.optimized = False
if 'results' not in st.session_state:
    st.session_state.results = None
# ---------------------- Load configurations End ----------------------

# ---------------------- Sidebar App Content Start ----------------------
# Title
st.title("Electricity Production Planning Optimizer")

# Collect plant data from sliders
plants = {}

st.sidebar.subheader("📂 Upload Custom Configurations")

# Expander for demand JSON instructions
with st.sidebar.expander("ℹ️ Demand JSON Instructions", expanded=True):
    st.write("Your demand JSON must be structured as follows:")
    st.code("""
{
  "periods": [
    {"name": "0h-6h", "hours": 6, "demand_mw": 15000},
    {"name": "6h-9h", "hours": 3, "demand_mw": 30000}
  ]
}
""", language="json")
    st.write("- `name`: period label (string)\n- `hours`: duration of period (int)\n- `demand_mw`: demand in MW (number)")

uploaded_demand = st.sidebar.file_uploader(
    "Upload demand JSON", type="json"
)

# Expander for plants JSON instructions
with st.sidebar.expander("ℹ️ Plants JSON Instructions", expanded=True):
    st.write("Your plants JSON must be structured as follows:")
    st.code("""
{
  "thermal_plants": [
    {
      "id": "A",
      "name": "Type A",
      "default": {"amount": 12, "p_min": 850, "p_max": 2000, "c_base": 1000, "c_mwh": 2.0, "c_start": 2000},
      "ranges": {"amount": [1, 20], "p_min": [100, 2000, 50], "p_max": [500, 5000, 50],
                 "c_base": [0, 5000, 50], "c_mwh": [0.0, 5.0, 0.1], "c_start": [0, 3000, 50]}
    }
  ],
  "hydro_plants": [
    {
      "id": "H1",
      "name": "Hydro 1",
      "default": {"amount": 1, "p_min": 900, "p_max": 900, "c_base": 90, "c_mwh": 0, "c_start": 1500, "v_abai": 0.31},
      "ranges": {"amount": [0, 2], "p_min": [500, 2000, 50], "p_max": [500, 2000, 50],
                 "c_base": [0, 500, 10], "c_start": [0, 2000, 50], "v_abai": [0.0, 1.0, 0.01]}
    }
  ]
}
""", language="json")
    st.write("- `thermal_plants` and `hydro_plants` must be lists of plant definitions.\n- Each plant requires `id`, `name`, `default`, and `ranges`.\n- `default` contains values used in optimization.\n- `ranges` define slider min, max, step.")
    
uploaded_plants = st.sidebar.file_uploader(
    "Upload plants JSON", type="json"
)

# Load configurations
demand_config, plants_config = load_config(uploaded_demand=uploaded_demand, uploaded_plants=uploaded_plants)

# Create demand dataframe
demand_df = pd.DataFrame([
    {"Period": p["name"], "Hours": p["hours"], "Demand (MW)": p["demand_mw"]}
    for p in demand_config["periods"]
])


st.sidebar.markdown("---")
st.sidebar.subheader("🏭 Thermal Power Plants")
for plant in plants_config["thermal_plants"]:
    with st.sidebar.expander(f"{plant['name']}", expanded=False):
        ranges = plant["ranges"]
        defaults = plant["default"]
        
        plants[plant["id"]] = {
            'amount': st.slider(
                f"Number available", 
                ranges["amount"][0], ranges["amount"][1], 
                defaults["amount"], 
                key=f"{plant['id']}_amount"
            ),
            'pMin': st.slider(
                f"Min Power (MW)", 
                ranges["p_min"][0], ranges["p_min"][1], 
                defaults["p_min"], ranges["p_min"][2], 
                key=f"{plant['id']}_pMin"
            ),
            'pMax': st.slider(
                f"Max Power (MW)", 
                ranges["p_max"][0], ranges["p_max"][1], 
                defaults["p_max"], ranges["p_max"][2], 
                key=f"{plant['id']}_pMax"
            ),
            'cBase': st.slider(
                f"Base Cost (€/h)", 
                ranges["c_base"][0], ranges["c_base"][1], 
                defaults["c_base"], ranges["c_base"][2], 
                key=f"{plant['id']}_cBase"
            ),
            'cMwh': st.slider(
                f"Variable Cost (€/MWh)", 
                float(ranges["c_mwh"][0]), float(ranges["c_mwh"][1]), 
                float(defaults["c_mwh"]), float(ranges["c_mwh"][2]), 
                key=f"{plant['id']}_cMwh"
            ),
            'cStart': st.slider(
                f"Startup Cost (€)", 
                ranges["c_start"][0], ranges["c_start"][1], 
                defaults["c_start"], ranges["c_start"][2], 
                key=f"{plant['id']}_cStart"
            ),
            'vAbai': 0
        }

st.sidebar.markdown("---")
st.sidebar.subheader("💧 Hydro Power Plants")
for plant in plants_config["hydro_plants"]:
    with st.sidebar.expander(f"{plant['name']}", expanded=False):
        ranges = plant["ranges"]
        defaults = plant["default"]
        
        plants[plant["id"]] = {
            'amount': st.slider(
                f"Number available", 
                ranges["amount"][0], ranges["amount"][1], 
                defaults["amount"], 
                key=f"{plant['id']}_amount"
            ),
            'pMin': st.slider(
                f"Fixed Power (MW)", 
                ranges["p_min"][0], ranges["p_min"][1], 
                defaults["p_min"], ranges["p_min"][2], 
                key=f"{plant['id']}_pMin"
            ),
            'pMax': st.slider(
                f"Max Power (MW)", 
                ranges["p_max"][0], ranges["p_max"][1], 
                defaults["p_max"], ranges["p_max"][2], 
                key=f"{plant['id']}_pMax"
            ),
            'cBase': st.slider(
                f"Base Cost (€/h)", 
                ranges["c_base"][0], ranges["c_base"][1], 
                defaults["c_base"], ranges["c_base"][2], 
                key=f"{plant['id']}_cBase"
            ),
            'cMwh': 0,
            'cStart': st.slider(
                f"Startup Cost (€)", 
                ranges["c_start"][0], ranges["c_start"][1], 
                defaults["c_start"], ranges["c_start"][2], 
                key=f"{plant['id']}_cStart"
            ),
            'vAbai': st.slider(
                f"Water Level Drop (m/h)", 
                float(ranges["v_abai"][0]), float(ranges["v_abai"][1]), 
                float(defaults["v_abai"]), float(ranges["v_abai"][2]), 
                key=f"{plant['id']}_vAbai"
            )
        }

st.sidebar.markdown("---")
# Additional parameters
st.sidebar.subheader("⚙️ Optimization Options")
reserve_pct = st.sidebar.slider("Reserve Requirement (%)", 0, 30, 15, 1) / 100
cyclic = st.sidebar.checkbox("Cyclic Planning (24h cycle)", value=True)
include_pumping = st.sidebar.checkbox("Include Pumping", value=True)


# Optimization button
if st.sidebar.button("🚀 Run Optimization", type="primary", use_container_width=True):
    with st.spinner("Optimizing... This may take a few seconds"):    
        optimize(demand_df, plants, reserve_pct, cyclic, include_pumping)
st.sidebar.markdown("---")

# ---------------------- Sidebar App Content End ----------------------


# ---------------------- Explanation Start ----------------------
st.header("Demand Profile")

col1, col2 = st.columns(2)

with col1:
    st.subheader("Demand Table")
    total_energy = sum(demand_df['Demand (MW)'] * demand_df['Hours'])
    st.metric("Total Energy Required", f"{total_energy:,.0f} MWh")
    st.dataframe(demand_df, hide_index=True, use_container_width=True)
    
with col2:
    st.subheader("Demand Chart")

    fig_demand = go.Figure()

    # Line + dots with values below
    fig_demand.add_trace(go.Scatter(
        x=demand_df['Period'],
        y=demand_df['Demand (MW)'],
        mode='lines+markers+text',
        marker=dict(color='red', size=8),
        line=dict(color='red', width=2),
        text = (demand_df['Demand (MW)']/1000).astype(str) + "k",
        textposition='middle left',  # Values below the dots
        showlegend=False
    ))

    fig_demand.update_layout(
        xaxis_title="Period",
        yaxis_title="Demand (MW)",
        height=350,
        plot_bgcolor='white'
    )

    st.plotly_chart(fig_demand, use_container_width=True)
# ---------------------- Explanation End ----------------------



# Main content
tab1, tab2 = st.tabs(["📐 Model Equations", "📊 Optimization Results"])


# ---------------------- Math Definitions Start ----------------------
with tab1:
    st.header("Mathematical Model")
    
    st.subheader("1️⃣ Objective Function")
    st.markdown("**Compact Form:**")
    st.latex(r"""
    \min \sum_{t=1}^{T} \sum_{p=1}^{P} \left[ (P_{t,p} - N_{t,p} \cdot P^{\min}_t) \cdot C^{MWh}_t \cdot H_p + N_{t,p} \cdot C^{base}_t \cdot H_p + X_{t,p} \cdot C^{start}_t \right]
    """)
    
    st.markdown("**Example with current values (Period 1, Plant A):**")
    example_plant = list(plants.keys())[0]
    p_data = plants[example_plant]
    h1 = demand_df.iloc[0]['Hours']
    st.latex(f"""
    (P_{{{example_plant},1}} - N_{{{example_plant},1}} \\cdot {p_data['pMin']}) \\cdot {p_data['cMwh']} \\cdot {h1} + N_{{{example_plant},1}} \\cdot {p_data['cBase']} \\cdot {h1} + X_{{{example_plant},1}} \\cdot {p_data['cStart']}
    """)
    
    st.markdown("**Variables:**")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("- $P_{t,p}$ : Power output (MW)")
        st.markdown("- $N_{t,p}$ : Active plants")
    with col2:
        st.markdown("- $X_{t,p}$ : Started plants")
        st.markdown("- $T$ : Number of plant types")
    with col3:
        st.markdown("- $P$ : Number of periods")
        st.markdown("- $H_p$ : Hours in period $p$")
    
    st.divider()
    st.subheader("2️⃣ Constraints")
    
    # Demand constraint
    st.markdown("**C1: Demand Satisfaction**")
    st.latex(r"\sum_{t=1}^{T} P_{t,p} \geq D_p \quad \forall p \in \{1, \ldots, P\}")
    st.markdown(f"**Example (Period 1):**")
    plant_sum = " + ".join([f"P_{{{name},1}}" for name in list(plants.keys())[:3]]) + " + \\ldots"
    st.latex(f"{plant_sum} \\geq {demand_df.iloc[0]['Demand (MW)']}")
    
    # Power limits
    st.markdown("**C2: Power Limits**")
    st.latex(r"N_{t,p} \cdot P^{\min}_t \leq P_{t,p} \leq N_{t,p} \cdot P^{\max}_t \quad \forall t, p")
    st.markdown(f"**Example (Plant {example_plant}, Period 1):**")
    st.latex(f"N_{{{example_plant},1}} \\cdot {p_data['pMin']} \\leq P_{{{example_plant},1}} \\leq N_{{{example_plant},1}} \\cdot {p_data['pMax']}")
    
    # Capacity
    st.markdown("**C3: Available Capacity**")
    st.latex(r"N_{t,p} \leq N^{\max}_t \quad \forall t, p")
    st.markdown(f"**Example (Plant {example_plant}):**")
    st.latex(f"N_{{{example_plant},p}} \\leq {p_data['amount']} \\quad \\forall p")
    
    # Startup logic
    st.markdown(f"**C4: Startup Logic {'(Cyclic)' if cyclic else '(Non-cyclic)'}**")
    st.latex(r"N_{t,p} \leq X_{t,p} + N_{t,p-1} \quad \forall t, p")
    if cyclic:
        st.markdown("*With $N_{t,0} = N_{t,P}$ for cyclic planning*")
    st.markdown(f"**Example (Plant {example_plant}, Period 2):**")
    st.latex(f"N_{{{example_plant},2}} \\leq X_{{{example_plant},2}} + N_{{{example_plant},1}}")
    
    # Reserve
    st.markdown(f"**C5: Reserve Requirement ({reserve_pct*100:.0f}%)**")
    st.latex(r"\sum_{t=1}^{T} (N_{t,p} \cdot P^{\max}_t - P_{t,p}) \geq \alpha \cdot D_p \quad \forall p")
    st.markdown(f"**Example (Period 1, $\\alpha = {reserve_pct}$):**")
    reserve_sum = " + ".join([f"(N_{{{name},1}} \\cdot {plants[name]['pMax']} - P_{{{name},1}})" for name in list(plants.keys())[:2]]) + " + \\ldots"
    st.latex(f"{reserve_sum} \\geq {reserve_pct} \\cdot {demand_df.iloc[0]['Demand (MW)']} = {reserve_pct * demand_df.iloc[0]['Demand (MW)']:.0f}")
    
    # Pumping
    if include_pumping:
        st.markdown("**C6: Pumping Constraint (Water Balance)**")
        st.latex(r"\sum_{p=1}^{P} \sum_{t=1}^{T} P_{t,p} \cdot H_p \geq \sum_{p=1}^{P} D_p \cdot H_p + \sum_{p=1}^{P} \sum_{t \in H} 3000 \cdot v_{t} \cdot N_{t,p} \cdot H_p")
        st.markdown("*Where $H$ is the set of hydro plants and $v_t$ is the water level drop rate (m/h)*")
        
        hydro_plants = [k for k in plants.keys() if plants[k]['vAbai'] > 0]
        if hydro_plants:
            st.markdown(f"**Example (Hydro plants: {', '.join(hydro_plants)}):**")
            water_example = " + ".join([f"3000 \\cdot {plants[h]['vAbai']:.2f} \\cdot N_{{{h},p}} \\cdot H_p" for h in hydro_plants[:2]])
            st.latex(f"\\text{{Total Production Energy}} \\geq \\text{{Total Demand Energy}} + ({water_example} + \\ldots)")
                 

# ---------------------- Math Definitions End ----------------------

# ---------------------- Optimization Results Start ----------------------
with tab2:
    if not st.session_state.optimized:
        st.info("👈 Configure parameters and click 'Run Optimization' to see results")
        
    elif st.session_state.results['status'] != 'Optimal':
        st.error("❌ **No Solution Found**")
        st.markdown("### Infeasible Constraints:")
        for constr in st.session_state.results.get('constraints', [])[:10]:
            st.markdown(f"- `{constr}`")
        st.markdown("**Suggestions:**")
        st.markdown("- Increase plant capacities or number of available plants")
        st.markdown("- Reduce reserve requirement")
        st.markdown("- Check if demand can be met with available capacity")
        
    else:
        results = st.session_state.results
        df = results['df']
        
        st.success(f"✅ **Optimal Solution Found!**")
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Cost", f"€{results['objective']/1000:.1f}k")
        col2.metric("Runtime", f"{results['runtime']:.3f}s")
        col3.metric("Gap", f"{results['gap']*100:.2f}%" if results['gap'] is not None else "-")
        col4.metric("Periods", len(demand_df))

        col5, col6, col7 = st.columns(3)
        col5.metric("Variables", results.get('num_vars', '-'))
        col6.metric("Constraints", results.get('num_constrs', '-'))
        col7.metric("Nodes Explored", results.get('nodes_explored', '-'))

        # Optionally, display model name below
        st.markdown(f"**Model Name:** {results.get('model_name', '-')}")
        st.markdown(f"**Solution Status:** {results.get('status', '-')}")
        st.divider()
        # Power production chart
        # Ensure proper ordering of periods
        period_order = demand_df["Period"].tolist()
        df['Period'] = pd.Categorical(df['Period'], categories=period_order, ordered=True)

        # Power production chart
        st.subheader("Power Production by Period")
        fig_power = go.Figure()

        for plant in df['Plant'].unique():
            plant_data = df[df['Plant'] == plant]
            # Sort by period to match categorical order
            plant_data = plant_data.sort_values('Period')
            fig_power.add_trace(go.Bar(
                name=plant,
                x=plant_data['Period'],
                y=plant_data['P_power'],
            ))

        # Add demand line
        demand_line = df.groupby('Period')['Demand'].first().reindex(period_order)
        fig_power.add_trace(go.Scatter(
            name='Demand',
            x=demand_line.index,
            y=demand_line.values,
            mode='lines+markers',
            line=dict(color='black', width=3, dash='dash'),
            marker=dict(size=10)
        ))

        fig_power.update_layout(
            barmode='stack',
            xaxis_title="Period",
            yaxis_title="Power (MW)",
            height=500,
            hovermode='x unified',
            plot_bgcolor='white'
        )
        st.plotly_chart(fig_power, use_container_width=True)

        
        # Active plants chart
        st.subheader("Active Plants by Period")
        fig_active = go.Figure()
        
        for plant in df['Plant'].unique():
            plant_data = df[df['Plant'] == plant]
            fig_active.add_trace(go.Bar(
                name=plant,
                x=plant_data['Period'],
                y=plant_data['N_active'],
            ))
        
        fig_active.update_layout(
            barmode='group',
            xaxis_title="Period",
            yaxis_title="Number of Active Plants",
            height=400,
            plot_bgcolor='white'
        )
        st.plotly_chart(fig_active, use_container_width=True)
        
        # Detailed results table
        st.subheader("Detailed Results")
        pivot_table = df.pivot_table(
            index='Period',
            columns='Plant',
            values=['P_power', 'N_active'],
            aggfunc='first'
        ).round(2)
        st.dataframe(pivot_table, use_container_width=True)
        
        # Download button
        # Create an Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Optimization Results')
            
            workbook = writer.book
            worksheet = writer.sheets['Optimization Results']
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Conditional formatting for P_power column
            max_row = worksheet.max_row
            col_idx = df.columns.get_loc("P_power") + 1  # openpyxl is 1-indexed
            color_rule = ColorScaleRule(
                start_type='min', start_color='FFFFFF',
                mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                end_type='max', end_color='63BE7B'
            )
            worksheet.conditional_formatting.add(f"{chr(64+col_idx)}2:{chr(64+col_idx)}{max_row}", color_rule)

        # Rewind the buffer
        output.seek(0)

        # Streamlit download button
        st.download_button(
            label="📥 Download Results (Excel)",
            data=output,
            file_name="optimization_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
# ---------------------- Optimization Results End ----------------------

footnote("cartoon_me.png", "José ALVES", "2025", "1.0.0", "https://jeduapf.github.io/")